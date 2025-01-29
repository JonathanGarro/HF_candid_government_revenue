#!/usr/bin/env python3

import pandas as pd
import requests
from openpyxl import load_workbook

API_KEY = "<API_TOKEN_HERE>"
API_URL = "https://api.candid.org/premier/v3"

def get_financial_data(ein, api_key):
	"""
	fetch financial data for a given EIN from the candid API
	depends on the "most_recent_year_financials dictionary within the return
	"""
	url = f"{API_URL}/{ein}"
	headers = {
		"accept": "application/json",
		"Subscription-Key": api_key
	}
	try:
		response = requests.get(url, headers=headers)
		response.raise_for_status()
		data = response.json()
		
		# access most recent financial data
		financials = data.get("data", {}).get("financials", {}).get("most_recent_year_financials", {})
		if not financials:
			print(f"No financial data available for EIN {ein}.")
			return None
		
		return {
			"revenue_contributions": financials.get("revenue_contributions"),
			"revenue_govt_grants": financials.get("revenue_govt_grants"),
			"revenue_total": financials.get("total_revenue")
		}
	except requests.exceptions.RequestException as e:
		print(f"Error fetching data for EIN {ein}: {e}")
		return None
	
def process_grantee_data(file_path, sheet_name, api_key):
	"""
	process the input excel file and write financial data back into the same file
	see code for the columns names you need to make sure the file includes 
	"""
	try:
		wb = load_workbook(file_path)
		if sheet_name not in wb.sheetnames:
			raise ValueError(f"Sheet {sheet_name} not found in the Excel file.")
			
		sheet = wb[sheet_name]
		df = pd.DataFrame(sheet.values)
		headers = df.iloc[0]
		df = pd.DataFrame(df.values[1:], columns=headers)
		
		if 'Organization: EIN' not in df.columns:
			raise ValueError("The sheet must contain a column named 'Organization: EIN'.")
			
		# ensure required columns exist, add any others you want to use
		if 'revenue_contributions' not in df.columns:
			df['revenue_contributions'] = None
		if 'revenue_govt_grants' not in df.columns:
			df['revenue_govt_grants'] = None
		if 'revenue_total' not in df.columns:
			df['revenue_total'] = None
			
		# process each EIN
		for index, row in df.iterrows():
			ein = row['Organization: EIN']
			
			if not ein or pd.isna(ein):
				print(f"Skipping row {index} due to missing EIN.")
				df.at[index, 'revenue_contributions'] = 'N/A'
				df.at[index, 'revenue_govt_grants'] = 'N/A'
				df.at[index, 'revenue_total'] = 'N/A'
				continue
			
			financial_data = get_financial_data(ein, api_key)
			
			if financial_data:
				df.at[index, 'revenue_contributions'] = financial_data.get('revenue_contributions', 'N/A')
				df.at[index, 'revenue_govt_grants'] = financial_data.get('revenue_govt_grants', 'N/A')
				df.at[index, 'revenue_total'] = financial_data.get('revenue_total', 'N/A')
			else:
				print(f"No data found for EIN {ein}. Writing N/A.")
				df.at[index, 'revenue_contributions'] = 'N/A'
				df.at[index, 'revenue_govt_grants'] = 'N/A'
				df.at[index, 'revenue_total'] = 'N/A'
				
		# save data back to input sheet
		for i, column in enumerate(df.columns, start=1):
			for j, value in enumerate(df[column], start=2):
				sheet.cell(row=j, column=i, value=value)
				
		wb.save(file_path)
		print(f"Data successfully updated in {file_path}.")
		
	except Exception as e:
		print(f"An error occurred: {e}")
		
if __name__ == "__main__":
	input_file = "grantee_data.xlsx"
	sheet_name = "candid_data"
	process_grantee_data(input_file, sheet_name, API_KEY)
	